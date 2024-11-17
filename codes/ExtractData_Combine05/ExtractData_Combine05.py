import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def combine_and_format_esg_data(file_llm, file_re, output_file='ESGDataOverall_Combine.xlsx',
                                confidence_value=0.5, adjust_value=0):
    # Load the Excel files
    df_llm = pd.read_excel(file_llm)
    df_re = pd.read_excel(file_re)

    # Check if RE DataFrame is empty
    if df_re.empty or df_re.shape[0] < 1:
        raise ValueError("RE DataFrame is empty or does not have enough data.")

    # Define new columns for the combined DataFrame
    new_columns = []
    for col in df_llm.columns:
        if '_value' in col or '_unit' in col:
            indicator = col.split('_')[0]
            if f'{indicator}_LLM_value' not in new_columns:
                new_columns.extend([f'{indicator}_LLM_value', f'{indicator}_LLM_unit',
                                    f'{indicator}_RE_value', f'{indicator}_RE_unit',
                                    f'{indicator}_combine_value', f'{indicator}_combine_unit',
                                    f'{indicator}_Diff', f'{indicator}_Inte'])
        else:
            new_columns.append(col)

    combined_df = pd.DataFrame(columns=new_columns)

    # Function to clean and convert values to numeric
    def clean_numeric(value):
        if isinstance(value, str):
            value = value.replace('$', '').replace(',', '')
        return pd.to_numeric(value, errors='coerce')

    # Function to compare values and return combined metrics
    def compare_values(indicator, llm_value, llm_unit, re_value, re_unit, confidence_value, adjust_value):
        llm_value = clean_numeric(llm_value)
        re_value = clean_numeric(re_value)

        qualitative_indicators = ['G-04', 'G-05', 'G-06', 'G-08']
        if indicator in qualitative_indicators:
            combine_unit = "TF"
            combine_value = 1 if not pd.isna(llm_value) or not pd.isna(re_value) else 0
            return (combine_value, combine_unit, None, 1)

        if pd.isna(llm_value) and pd.isna(re_value):
            return (None, None, None, -1)
        elif pd.isna(llm_value) or pd.isna(re_value):
            combine_value = llm_value if not pd.isna(llm_value) else re_value
            combine_unit = llm_unit if not pd.isna(llm_unit) else re_unit
            return (combine_value, combine_unit, None, 0 if combine_unit is None else 1)
        else:
            diff = abs(llm_value - re_value) / max(llm_value, re_value)
            if diff < confidence_value:
                combine_value = (llm_value + re_value) * (0.5 + adjust_value)
                combine_unit = llm_unit if abs(llm_value - combine_value) < abs(re_value - combine_value) else re_unit
            else:
                combine_value = max(llm_value, re_value)
                combine_unit = llm_unit if llm_value == combine_value else re_unit
            return (combine_value, combine_unit, diff, 1)

    # Process each row
    for idx_llm, row_llm in df_llm.iterrows():
        matched_row = df_re.iloc[0]

        new_row = {col: row_llm[col] for col in df_llm.columns if col not in ['_value', '_unit']}

        for col in df_llm.columns:
            if '_value' in col:
                indicator = col.split('_')[0]
                llm_value = row_llm[f'{indicator}_value']
                llm_unit = row_llm[f'{indicator}_unit']
                re_value = matched_row[f'{indicator}_value']
                re_unit = matched_row[f'{indicator}_unit']

                combine_value, combine_unit, diff, inte = compare_values(
                    indicator, llm_value, llm_unit, re_value, re_unit,
                    confidence_value=confidence_value,
                    adjust_value=adjust_value
                )

                new_row[f'{indicator}_LLM_value'] = llm_value
                new_row[f'{indicator}_LLM_unit'] = llm_unit
                new_row[f'{indicator}_RE_value'] = re_value
                new_row[f'{indicator}_RE_unit'] = re_unit
                new_row[f'{indicator}_combine_value'] = combine_value
                new_row[f'{indicator}_combine_unit'] = combine_unit
                new_row[f'{indicator}_Diff'] = diff
                new_row[f'{indicator}_Inte'] = inte

        for col in combined_df.columns:
            if col not in new_row:
                new_row[col] = np.nan

        new_row_df = pd.DataFrame([new_row])[combined_df.columns]
        combined_df = pd.concat([combined_df, new_row_df], ignore_index=True)

    combined_df.to_excel(output_file, index=False)

    # Apply conditional formatting
    wb = load_workbook(output_file)
    ws = wb.active

    light_gray_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    light_orange_fill = PatternFill(start_color='FFD580', end_color='FFD580', fill_type='solid')

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if cell.value is None or cell.value == '':
                cell.fill = light_gray_fill
            elif 'Diff' in ws.cell(row=1, column=cell.column).value:
                if isinstance(cell.value, (int, float)) and cell.value > confidence_value:
                    cell.fill = orange_fill
                else:
                    cell.fill = light_green_fill
            elif 'Inte' in ws.cell(row=1, column=cell.column).value:
                if cell.value == -1:
                    cell.fill = orange_fill
                elif cell.value == 0:
                    cell.fill = light_orange_fill
                elif cell.value == 1:
                    cell.fill = light_green_fill

    wb.save(output_file)
    return output_file


