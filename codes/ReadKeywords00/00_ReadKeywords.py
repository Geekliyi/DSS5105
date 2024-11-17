import openpyxl
import csv

def read_keywords(a_table_path, b_table_path, c_table_path):
    # Open A table (xlsx file)
    wb_a = openpyxl.load_workbook(a_table_path)
    sheet_a = wb_a.active

    # Read B and C tables (CSV files)
    with open(b_table_path, mode='r', newline='', encoding='utf-8') as file_b:
        reader_b = list(csv.reader(file_b))

    with open(c_table_path, mode='r', newline='', encoding='utf-8') as file_c:
        reader_c = list(csv.reader(file_c))

    # Get the number of reports and metrics
    report_count = sheet_a.max_row - 2  # Number of reports, starting from row 3
    metric_count = (sheet_a.max_column - 8) // 4  # Number of metrics per report

    # Resize the B and C lists to match the metrics if needed
    while len(reader_b) < metric_count + 1:
        reader_b.append([len(reader_b)])  # Add new rows with metric index
    while len(reader_c) < metric_count + 1:
        reader_c.append([len(reader_c)])  # Add new rows with metric index

    # Iterate through each report, starting from row 3
    for report_idx in range(3, report_count + 3):  # Start from row 3
        # Iterate through each metric in the report
        for metric_idx in range(metric_count):
            value_col = 10 + metric_idx * 4  # Column for value_keyword (2nd column of the metric)
            unit_col = 12 + metric_idx * 4  # Column for unit_keyword (4th column of the metric)

            value_keyword = sheet_a.cell(row=report_idx, column=value_col).value
            unit_keyword = sheet_a.cell(row=report_idx, column=unit_col).value
            print(value_keyword, unit_keyword)

            # Store value_keyword in the B list
            while len(reader_b[metric_idx + 1]) < report_idx - 1:  # Expand if necessary
                reader_b[metric_idx + 1].append('')
            reader_b[metric_idx + 1].append(value_keyword or '')

            # Store unit_keyword in the C list
            while len(reader_c[metric_idx + 1]) < report_idx - 1:  # Expand if necessary
                reader_c[metric_idx + 1].append('')
            reader_c[metric_idx + 1].append(unit_keyword or '')

    # Remove duplicates and empty values from B and C
    for row in range(1, metric_count + 1):
        # Process B table
        b_values = list(filter(None, set(reader_b[row][1:])))  # Remove duplicates and empty
        reader_b[row] = [row] + b_values

        # Process C table
        c_values = list(filter(None, set(reader_c[row][1:])))  # Remove duplicates and empty
        reader_c[row] = [row] + c_values

    # Append the updated data back to B and C CSV files
    with open(b_table_path, mode='a', newline='', encoding='utf-8') as file_b:
        writer_b = csv.writer(file_b)
        # Only write if there's data to append
        for row in reader_b[1:]:  # Skip header row
            if len(row) > 1:  # Only write if there's data beyond the index
                writer_b.writerow(row)

    with open(c_table_path, mode='a', newline='', encoding='utf-8') as file_c:
        writer_c = csv.writer(file_c)
        # Only write if there's data to append
        for row in reader_c[1:]:  # Skip header row
            if len(row) > 1:  # Only write if there's data beyond the index
                writer_c.writerow(row)

read_keywords('ESGDataOverall_sample.xlsx', 'ESGValueKeywordsDic.csv', 'ESGUnitKeywordsDic.csv')